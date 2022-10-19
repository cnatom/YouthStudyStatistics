from openpyxl import load_workbook, Workbook
from collections import Counter
from openpyxl.worksheet.worksheet import Worksheet


def get_counter(workbook: str) -> dict:
    result: dict = {}
    raw: Workbook = load_workbook(workbook)
    for sheet_name in raw.sheetnames:
        sheet: Worksheet = raw[sheet_name]
        cols: list = sheet["B"][1:]
        cols = list(filter(None, [item.value for item in cols]))
        cur_result: dict = Counter(cols)
        result[sheet_name] = cur_result
    return result


def export(filename: str, counter: dict):
    total_book: Workbook = load_workbook("total.xlsx")
    for key, value_dict in counter.items():
        summary_total, summary_finished = 0, 0
        '''遍历未完成的年级，key是年级，value_dict是年级内各个班的未完成数量'''
        sheet: Worksheet = total_book[key]
        sheet.cell(row=1, column=3, value="已完成")
        sheet.cell(row=1, column=4, value="完成率")
        for i in range(2, sheet.max_row + 1):
            '''遍历total.xlsx中某个年级Sheet的每一行'''
            cell_key: str = sheet.cell(row=i, column=1).value
            total: int = sheet.cell(row=i, column=2).value  # 当前班级总人数
            if cell_key in value_dict.keys():
                unfinished: int = int(value_dict[cell_key])  # 当前班级未完成人数
                percent: str = "%.2f%%" % ((total - unfinished) / total * 100)  # 当前班级完成率
                sheet.cell(row=i, column=3, value=total - unfinished)
                sheet.cell(row=i, column=4, value=percent)
                summary_finished = summary_finished + total - unfinished
            else:
                sheet.cell(row=i, column=3, value=0)
                sheet.cell(row=i, column=4, value="00.00%")
            summary_total = summary_total + total

        sheet.cell(row=sheet.max_row + 1, column=1, value="汇总")
        sheet.cell(row=sheet.max_row, column=2, value=summary_total)
        sheet.cell(row=sheet.max_row, column=3, value=summary_finished)
        sheet.cell(row=sheet.max_row, column=4, value="%.2f%%" % (summary_finished / summary_total * 100))

    total_book.save(filename)


if __name__ == '__main__':
    counter: dict = get_counter(workbook="unfinished.xlsx")
    print(counter)
    export(filename="result.xlsx", counter=counter)
